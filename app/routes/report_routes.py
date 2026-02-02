from fastapi import APIRouter, Body, Response, HTTPException, Depends
from fastapi.responses import StreamingResponse
from db import fetch_query
from pydantic import BaseModel
from typing import Optional
from datetime import date, datetime
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import requests
import os
import logging
from auth import get_current_user, get_customer_id

# Configure logging
logger = logging.getLogger(__name__)

router = APIRouter()


def get_weather_api_key():
    """Fetch OpenWeather API key from database"""
    try:
        result = fetch_query(
            "SELECT key_value FROM api_keys WHERE key_name = 'openweather_api_key' AND user_id IS NULL LIMIT 1"
        )
        if result and result[0].get('key_value'):
            logger.info(f"Weather API key found in database")
            return result[0]['key_value']
        logger.warning("Weather API key not found in database")
        return None
    except Exception as e:
        logger.error(f"Failed to fetch weather API key: {str(e)}")
        return None

class ReportFilters(BaseModel):
    """Filters for generating reports"""
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    property_id: Optional[int] = None
    user_id: Optional[int] = None
    contractor_name: Optional[str] = None
    equipment: Optional[str] = None
    winter_event_id: Optional[int] = None

@router.post("/report/by-product/")
def report_by_product(filters: ReportFilters, customer_id: str = Depends(get_customer_id)):
    start = filters.start_date
    end = filters.end_date
    property_id = filters.property_id
    user_id = filters.user_id

    where_clauses = ["wl.customer_id = %s"]
    params = [customer_id]

    if start and end:
        where_clauses.append("(gs.service_date BETWEEN %s AND %s OR wl.time_in BETWEEN %s AND %s)")
        params += [start, end, start, end]

    if property_id:
        where_clauses.append("(gs.property_id = %s OR wl.property_id = %s)")
        params += [property_id, property_id]

    if user_id:
        where_clauses.append("(gs.worker_name = (SELECT name FROM users WHERE id = %s AND customer_id = %s) OR wl.worker_name = (SELECT name FROM users WHERE id = %s AND customer_id = %s))")
        params += [user_id, customer_id, user_id, customer_id]

    where_sql = " AND ".join(where_clauses)
    if where_sql:
        where_sql = "WHERE " + where_sql

    query = f"""
        SELECT product_name, SUM(quantity) AS total_used FROM (
            SELECT lp.name AS product_name, gs.product_quantity AS quantity
            FROM green_services gs
            JOIN landscape_products lp ON gs.product_used = lp.id
            {where_sql}
            UNION ALL
            SELECT 'Bulk Salt', wl.bulk_salt_qty FROM winter_ops_logs wl {where_sql}
            UNION ALL
            SELECT 'Bag Salt', wl.bag_salt_qty FROM winter_ops_logs wl {where_sql}
            UNION ALL
            SELECT 'Calcium Chloride', wl.calcium_chloride_qty FROM winter_ops_logs wl {where_sql}
        ) AS combined
        GROUP BY product_name
        ORDER BY total_used DESC;
    """
    return fetch_query(query, params)

@router.post("/report/by-property/")
def report_by_property(filters: ReportFilters, customer_id: str = Depends(get_customer_id)):
    start = filters.start_date
    end = filters.end_date
    property_id = filters.property_id
    user_id = filters.user_id

    conditions = ["l.customer_id = %s"]
    params = [customer_id]

    if start and end:
        conditions.append("(w.time_in BETWEEN %s AND %s OR g.time_in BETWEEN %s AND %s)")
        params += [start, end, start, end]
    if property_id:
        conditions.append("l.id = %s")
        params.append(property_id)
    if user_id:
        conditions.append("(w.worker_name = (SELECT name FROM users WHERE id = %s AND customer_id = %s) OR g.worker_name = (SELECT name FROM users WHERE id = %s AND customer_id = %s))")
        params += [user_id, customer_id, user_id, customer_id]

    where_clause = " AND ".join(conditions)
    if where_clause:
        where_clause = "WHERE " + where_clause

    query = f"""
        SELECT
            l.name AS property,
            COUNT(DISTINCT w.id) AS winter_logs,
            COUNT(DISTINCT g.id) AS green_logs,
            SUM(TIMESTAMPDIFF(SECOND, w.time_in, w.time_out) / 3600) AS winter_hours,
            SUM(TIMESTAMPDIFF(SECOND, g.time_in, g.time_out) / 3600) AS green_hours,
            SUM(w.bulk_salt_qty) AS bulk_salt,
            SUM(w.bag_salt_qty) AS bag_salt,
            SUM(w.calcium_chloride_qty) AS calcium_chloride,
            SUM(g.quantity_used) AS green_products_used
        FROM locations l
        LEFT JOIN winter_ops_logs w ON w.property_id = l.id AND w.customer_id = %s
        LEFT JOIN green_services_logs g ON g.property_id = l.id
        {where_clause}
        GROUP BY l.name
        ORDER BY l.name;
    """
    params.insert(1, customer_id)  # Add customer_id for the JOIN condition
    return fetch_query(query, params)

@router.post("/report/by-user/")
def report_by_user(filters: ReportFilters, customer_id: str = Depends(get_customer_id)):
    start = filters.start_date
    end = filters.end_date
    property_id = filters.property_id
    user_id = filters.user_id

    conditions = ["u.customer_id = %s"]
    params = [customer_id]

    if start and end:
        conditions.append("(w.time_in BETWEEN %s AND %s OR g.time_in BETWEEN %s AND %s)")
        params += [start, end, start, end]
    if property_id:
        conditions.append("(w.property_id = %s OR g.property_id = %s)")
        params += [property_id, property_id]
    if user_id:
        conditions.append("(w.worker_name = (SELECT name FROM users WHERE id = %s AND customer_id = %s) OR g.worker_name = (SELECT name FROM users WHERE id = %s AND customer_id = %s))")
        params += [user_id, customer_id, user_id, customer_id]

    where_clause = " AND ".join(conditions)
    if where_clause:
        where_clause = "WHERE " + where_clause

    query = f"""
        SELECT
            COALESCE(w.worker_name, g.worker_name) AS subcontractor,
            COUNT(DISTINCT w.id) AS winter_logs,
            COUNT(DISTINCT g.id) AS green_logs,
            SUM(TIMESTAMPDIFF(SECOND, w.time_in, w.time_out) / 3600) AS winter_hours,
            SUM(TIMESTAMPDIFF(SECOND, g.time_in, g.time_out) / 3600) AS green_hours,
            SUM(w.bulk_salt_qty) AS bulk_salt,
            SUM(w.bag_salt_qty) AS bag_salt,
            SUM(w.calcium_chloride_qty) AS calcium_chloride,
            SUM(g.quantity_used) AS green_products_used
        FROM users u
        LEFT JOIN winter_ops_logs w ON w.worker_name = u.name AND w.customer_id = %s
        LEFT JOIN green_services_logs g ON g.worker_name = u.name
        {where_clause}
        GROUP BY subcontractor
        ORDER BY subcontractor;
    """
    params.insert(1, customer_id)  # Add customer_id for the JOIN condition
    return fetch_query(query, params)

@router.post("/export/contractor-timesheets/")
def export_contractor_timesheets(filters: ReportFilters, customer_id: str = Depends(get_customer_id)):
    """
    Export contractor timesheets in Excel format
    Similar to the format: Date/Contractor name with timesheet details
    """
    start = filters.start_date
    end = filters.end_date

    # Fetch winter logs
    where_parts = ["w.customer_id = %s"]
    params = [customer_id]

    if start and end:
        where_parts.append("time_in BETWEEN %s AND %s")
        params += [start, end]

    where_clause = "WHERE " + " AND ".join(where_parts) if where_parts else ""

    query = f"""
        SELECT
            w.id,
            w.worker_name,
            DATE(w.time_in) as work_date,
            w.time_in,
            w.time_out,
            TIMESTAMPDIFF(MINUTE, w.time_in, w.time_out) as total_minutes,
            TIMESTAMPDIFF(SECOND, w.time_in, w.time_out) / 3600 as hours,
            l.name as site,
            w.equipment,
            w.bulk_salt_qty,
            w.bag_salt_qty,
            w.calcium_chloride_qty,
            w.notes
        FROM winter_ops_logs w
        JOIN locations l ON w.property_id = l.id AND l.customer_id = %s
        {where_clause}
        ORDER BY w.worker_name, w.time_in
    """
    params.insert(1, customer_id)  # Add customer_id for the JOIN condition

    logs = fetch_query(query, params if params else None)

    if not logs:
        raise HTTPException(status_code=404, detail="No logs found for the specified filters")

    # Create Excel workbook with pandas
    output = BytesIO()

    # Group by contractor and date
    df = pd.DataFrame(logs)

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for contractor in df['worker_name'].unique():
            contractor_logs = df[df['worker_name'] == contractor]

            for work_date in contractor_logs['work_date'].unique():
                date_logs = contractor_logs[contractor_logs['work_date'] == work_date]

                # Calculate totals
                total_hours = date_logs['hours'].sum()
                total_minutes = date_logs['total_minutes'].sum()

                # Format sheet name (date_contractor)
                sheet_name = f"{work_date.strftime('%m%d%Y')}_{contractor}"[:31]  # Excel limit

                # Create formatted dataframe
                export_data = []
                export_data.append(['Contractor:', contractor, '', '', '', f'Total Time: {total_hours:.2f} hrs'])
                export_data.append(['Start Time', 'End Time', 'Total Min', 'HRS', 'Site', 'Equipment', 'Qty Salt (yrd)'])

                for _, row in date_logs.iterrows():
                    start_time = pd.to_datetime(row['time_in']).strftime('%H%M')
                    end_time = pd.to_datetime(row['time_out']).strftime('%H%M')
                    export_data.append([
                        start_time,
                        end_time,
                        int(row['total_minutes']),
                        round(row['hours'], 2),
                        row['site'],
                        row['equipment'] if row['equipment'] else '',
                        row['bulk_salt_qty'] if row['bulk_salt_qty'] else ''
                    ])

                # Create DataFrame and write to sheet
                sheet_df = pd.DataFrame(export_data)
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

    output.seek(0)

    # Generate filename
    filename = f"contractor_timesheets_{start}_{end}.xlsx" if start and end else "contractor_timesheets.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/export/property-logs/")
def export_property_logs(filters: ReportFilters, customer_id: str = Depends(get_customer_id)):
    """
    Export property-based logs in Excel format
    One sheet per property with all contractor visits
    """
    start = filters.start_date
    end = filters.end_date
    property_id = filters.property_id

    where_parts = ["w.customer_id = %s"]
    params = [customer_id]

    if start and end:
        where_parts.append("time_in BETWEEN %s AND %s")
        params += [start, end]

    if property_id:
        where_parts.append("w.property_id = %s")
        params.append(property_id)

    where_clause = "WHERE " + " AND ".join(where_parts) if where_parts else ""

    query = f"""
        SELECT
            l.name as property_name,
            w.worker_name,
            DATE(w.time_in) as work_date,
            w.time_in,
            w.time_out,
            TIMESTAMPDIFF(MINUTE, w.time_in, w.time_out) as total_minutes,
            TIMESTAMPDIFF(SECOND, w.time_in, w.time_out) / 3600 as hours,
            w.equipment,
            w.bulk_salt_qty,
            w.bag_salt_qty,
            w.calcium_chloride_qty,
            w.notes
        FROM winter_ops_logs w
        JOIN locations l ON w.property_id = l.id AND l.customer_id = %s
        {where_clause}
        ORDER BY l.name, w.time_in
    """
    params.insert(1, customer_id)  # Add customer_id for the JOIN condition

    logs = fetch_query(query, params if params else None)

    if not logs:
        raise HTTPException(status_code=404, detail="No logs found for the specified filters")

    output = BytesIO()
    df = pd.DataFrame(logs)

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for property_name in df['property_name'].unique():
            property_logs = df[df['property_name'] == property_name]

            # Calculate totals
            total_hours = property_logs['hours'].sum()
            total_salt = property_logs['bulk_salt_qty'].sum()

            # Format sheet name
            sheet_name = property_name[:31]  # Excel limit

            # Create formatted dataframe
            export_data = []
            export_data.append([f'Property: {property_name}', '', '', '', '', f'Total Hours: {total_hours:.2f}', f'Total Salt: {total_salt:.2f} yrd'])
            export_data.append(['Date', 'Contractor', 'Equipment', 'Start Time', 'End Time', 'Hours', 'Bulk Salt (yrd)', 'Bag Salt', 'Calcium', 'Notes'])

            for _, row in property_logs.iterrows():
                work_date = pd.to_datetime(row['work_date']).strftime('%m/%d/%Y')
                start_time = pd.to_datetime(row['time_in']).strftime('%H:%M')
                end_time = pd.to_datetime(row['time_out']).strftime('%H:%M')
                export_data.append([
                    work_date,
                    row['worker_name'],
                    row['equipment'] if row['equipment'] else '',
                    start_time,
                    end_time,
                    round(row['hours'], 2),
                    row['bulk_salt_qty'] if row['bulk_salt_qty'] else '',
                    row['bag_salt_qty'] if row['bag_salt_qty'] else '',
                    row['calcium_chloride_qty'] if row['calcium_chloride_qty'] else '',
                    row['notes'] if row['notes'] else ''
                ])

            sheet_df = pd.DataFrame(export_data)
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

    output.seek(0)

    filename = f"property_logs_{start}_{end}.xlsx" if start and end else "property_logs.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/export/billing-report/")
def export_billing_report(filters: ReportFilters, customer_id: str = Depends(get_customer_id)):
    """
    Export billing report with ONE SHEET PER CONTRACTOR
    Format per sheet:
    - Top: Contractor name (e.g., "AgriFarms")
    - Columns: Date, Time In, Time Out, Total Min, HR $, Site, Qty Salt (yrd)
    - Bottom: Total hours for that contractor
    """
    start = filters.start_date
    end = filters.end_date
    contractor_name = filters.contractor_name
    equipment = filters.equipment

    where_parts = ["w.customer_id = %s"]
    params = [customer_id]

    if start and end:
        where_parts.append("DATE(w.time_in) BETWEEN %s AND %s")
        params += [start, end]

    if contractor_name:
        where_parts.append("w.contractor_name = %s")
        params.append(contractor_name)

    if equipment:
        where_parts.append("w.equipment = %s")
        params.append(equipment)

    where_clause = "WHERE " + " AND ".join(where_parts) if where_parts else ""

    query = f"""
        SELECT
            DATE(w.time_in) as work_date,
            w.time_in,
            w.time_out,
            TIMESTAMPDIFF(MINUTE, w.time_in, w.time_out) as total_minutes,
            TIMESTAMPDIFF(SECOND, w.time_in, w.time_out) / 3600 as hours,
            l.name as site,
            w.bulk_salt_qty,
            w.contractor_name,
            w.equipment
        FROM winter_ops_logs w
        JOIN locations l ON w.property_id = l.id AND l.customer_id = %s
        {where_clause}
        ORDER BY w.contractor_name, w.time_in
    """
    params.insert(1, customer_id)  # Add customer_id for the JOIN condition

    logs = fetch_query(query, params if params else None)

    if not logs:
        raise HTTPException(status_code=404, detail="No logs found for the specified filters")

    output = BytesIO()
    df = pd.DataFrame(logs)
    truck_rate = 150  # $150/hr

    # Group by contractor and create one sheet per contractor
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for contractor in df['contractor_name'].unique():
            contractor_logs = df[df['contractor_name'] == contractor]

            # Calculate contractor totals
            contractor_total_minutes = contractor_logs['total_minutes'].sum()
            contractor_total_hours = contractor_total_minutes / 60

            # Create sheet data
            export_data = []

            # Header: Contractor name
            export_data.append([contractor])
            export_data.append([])  # Blank row

            # Column headers
            export_data.append(['Date', 'Start Time', 'End Time', 'Total Min', 'HR $', 'Site', 'Qty Salt (yrd)'])

            # Data rows
            for _, row in contractor_logs.iterrows():
                work_date = pd.to_datetime(row['work_date']).strftime('%m/%d/%Y')
                start_time = pd.to_datetime(row['time_in']).strftime('%I:%M:%S %p')
                end_time = pd.to_datetime(row['time_out']).strftime('%I:%M:%S %p') if row['time_out'] else ''
                total_min = int(row['total_minutes'] or 0)
                hours = float(row['hours'] or 0)
                hr_cost = round(hours * truck_rate, 2)
                site = row['site']
                salt_qty = row['bulk_salt_qty'] if row['bulk_salt_qty'] else ''

                export_data.append([
                    work_date,
                    start_time,
                    end_time,
                    total_min,
                    hr_cost,
                    site,
                    salt_qty
                ])

            # Add totals row at bottom
            export_data.append([])  # Blank row
            export_data.append(['Total Hours:', f'{contractor_total_hours:.2f}'])

            # Create DataFrame for this contractor
            sheet_df = pd.DataFrame(export_data)

            # Sheet name (sanitize contractor name for Excel)
            sheet_name = contractor[:31]  # Excel 31 char limit

            # Write to sheet
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

            # Auto-adjust column widths
            worksheet = writer.sheets[sheet_name]
            for idx in range(len(sheet_df.columns)):
                max_length = max(
                    sheet_df.iloc[:, idx].astype(str).map(len).max(),
                    10
                )
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 2, 30)

    output.seek(0)

    contractor_suffix = f"_{contractor_name}" if contractor_name else "_all_contractors"
    filename = f"billing_report{contractor_suffix}_{start}_{end}.xlsx" if start and end else f"billing_report{contractor_suffix}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/export/winter-logs/")
def export_winter_logs(filters: ReportFilters):
    """
    Export winter operations logs in Excel format
    Comprehensive export with all log details for the ViewWinterLogs page
    """
    start = filters.start_date
    end = filters.end_date
    property_id = filters.property_id
    contractor_name = filters.contractor_name
    equipment = filters.equipment

    where_parts = []
    params = []

    if start and end:
        where_parts.append("DATE(w.time_in) BETWEEN %s AND %s")
        params += [start, end]

    if property_id:
        where_parts.append("w.property_id = %s")
        params.append(property_id)

    if contractor_name:
        where_parts.append("w.contractor_name = %s")
        params.append(contractor_name)

    if equipment:
        where_parts.append("w.equipment = %s")
        params.append(equipment)

    where_clause = "WHERE " + " AND ".join(where_parts) if where_parts else ""

    query = f"""
        SELECT
            DATE(w.time_in) as work_date,
            l.name as property_name,
            w.contractor_name,
            w.worker_name,
            w.equipment,
            w.time_in,
            w.time_out,
            TIMESTAMPDIFF(SECOND, w.time_in, w.time_out) / 3600 as hours,
            w.bulk_salt_qty,
            w.bag_salt_qty,
            w.calcium_chloride_qty,
            w.customer_provided,
            w.notes,
            we.event_name as winter_event_name
        FROM winter_ops_logs w
        JOIN locations l ON w.property_id = l.id
        LEFT JOIN winter_events we ON w.winter_event_id = we.id
        {where_clause}
        ORDER BY w.time_in DESC
    """

    logs = fetch_query(query, params if params else None)

    if not logs:
        raise HTTPException(status_code=404, detail="No logs found for the specified filters")

    output = BytesIO()

    # Create comprehensive Excel export
    export_data = []

    # Header row
    export_data.append([
        'Date', 'Property', 'Contractor', 'Worker', 'Equipment',
        'Time In', 'Time Out', 'Hours', 'Bulk Salt (tons)', 'Bag Salt',
        'Calcium Chloride (lbs)', 'Customer Provided', 'Notes', 'Winter Event'
    ])

    # Calculate totals
    total_hours = 0
    total_bulk_salt = 0
    total_bag_salt = 0
    total_calcium = 0

    # Data rows
    for row in logs:
        work_date = pd.to_datetime(row['work_date']).strftime('%m/%d/%Y')
        time_in = pd.to_datetime(row['time_in']).strftime('%m/%d/%Y %H:%M')
        time_out = pd.to_datetime(row['time_out']).strftime('%m/%d/%Y %H:%M') if row['time_out'] else ''
        hours = round(row['hours'], 2) if row['hours'] else 0

        total_hours += hours
        total_bulk_salt += float(row['bulk_salt_qty'] or 0)
        total_bag_salt += float(row['bag_salt_qty'] or 0)
        total_calcium += float(row['calcium_chloride_qty'] or 0)

        export_data.append([
            work_date,
            row['property_name'],
            row['contractor_name'] or '',
            row['worker_name'] or '',
            row['equipment'] or '',
            time_in,
            time_out,
            hours,
            row['bulk_salt_qty'] if row['bulk_salt_qty'] else '',
            row['bag_salt_qty'] if row['bag_salt_qty'] else '',
            row['calcium_chloride_qty'] if row['calcium_chloride_qty'] else '',
            'Yes' if row['customer_provided'] else 'No',
            row['notes'] if row['notes'] else '',
            row['winter_event_name'] if row['winter_event_name'] else ''
        ])

    # Add totals row
    export_data.append([])
    export_data.append([
        'TOTALS', f'{len(logs)} logs', '', '', '', '', '',
        f'{total_hours:.2f}', f'{total_bulk_salt:.2f}', f'{total_bag_salt}',
        f'{total_calcium:.2f}', '', '', ''
    ])

    # Create DataFrame and export
    df = pd.DataFrame(export_data)

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Winter Logs', index=False, header=False)

        # Auto-adjust column widths
        worksheet = writer.sheets['Winter Logs']
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(str(col))
            )
            worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 2, 50)

    output.seek(0)

    filename = f"winter_logs_{start}_{end}.xlsx" if start and end else "winter_logs.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/export/field-service-reports/")
def export_field_service_reports(filters: ReportFilters):
    """
    Export field service reports grouped by Property + Equipment
    One sheet per property-equipment combination with weather data
    Format: AgriFarms template with all contractors for that equipment
    Filters: winter_event_id OR (start_date AND end_date), plus optional property_id
    If no filters provided, exports ALL properties and equipment for ALL time
    """
    start = filters.start_date
    end = filters.end_date
    property_id = filters.property_id
    winter_event_id = filters.winter_event_id

    where_parts = []
    params = []

    # Priority: winter event ID overrides date range
    if winter_event_id:
        where_parts.append("w.winter_event_id = %s")
        params.append(winter_event_id)
    elif start and end:
        where_parts.append("DATE(w.time_in) BETWEEN %s AND %s")
        params += [start, end]

    if property_id:
        where_parts.append("w.property_id = %s")
        params.append(property_id)

    where_clause = "WHERE " + " AND ".join(where_parts) if where_parts else ""

    # Get event name if filtering by event
    event_name = None
    if winter_event_id:
        event_query = "SELECT event_name FROM winter_events WHERE id = %s"
        event_result = fetch_query(event_query, (winter_event_id,))
        if event_result and len(event_result) > 0:
            event_name = event_result[0]['event_name']

    # Get all logs grouped by property and equipment
    query = f"""
        SELECT
            l.id as property_id,
            l.name as property_name,
            l.address as property_address,
            l.area_manager,
            l.latitude,
            l.longitude,
            w.equipment,
            w.time_in,
            w.time_out,
            TIMESTAMPDIFF(SECOND, w.time_in, w.time_out) / 3600 as hours,
            w.contractor_name,
            w.worker_name,
            w.notes
        FROM winter_ops_logs w
        JOIN locations l ON w.property_id = l.id
        {where_clause}
        ORDER BY l.name, w.equipment, w.time_in
    """

    logs = fetch_query(query, params if params else None)

    if not logs:
        raise HTTPException(status_code=404, detail="No logs found for the specified filters")

    df = pd.DataFrame(logs)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Group by property + equipment
    for (prop_name, equipment), group in df.groupby(['property_name', 'equipment']):
        # Get property details from first row
        first_row = group.iloc[0]

        # Sanitize sheet name
        sheet_name = f"{prop_name[:15]}-{equipment[:15]}"[:31]

        ws = wb.create_sheet(sheet_name)

        # AgriFarms header
        ws['A1'] = 'AgriFarms LLC'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')

        # Property Information
        ws['A6'] = 'Property Information'
        ws['A6'].font = Font(bold=True, size=12)
        ws['A6'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        ws['A7'] = 'Property Name:'
        ws['B7'] = prop_name
        ws['B7'].font = Font(bold=True)

        ws['A8'] = 'Address:'
        ws['B8'] = first_row['property_address'] or ''

        ws['A9'] = 'Date of Service:'
        if event_name:
            ws['B9'] = f"Event: {event_name}"
        elif start and end:
            ws['B9'] = f"{start} to {end}"
        else:
            ws['B9'] = 'All Time'

        ws['A10'] = 'Service Window:'
        ws['B10'] = 'Snow Event Response'

        ws['A11'] = 'Area Manager:'
        ws['B11'] = first_row['area_manager'] or ''

        # Equipment
        ws['A13'] = 'Equipment Used:'
        ws['B13'] = equipment
        ws['B13'].font = Font(bold=True)

        # Time Log Headers
        row = 16
        headers = ['Date', 'Time In', 'Time Out', 'Total Hours', 'Contractor', 'Worker Name', 'Service Performed', 'Notes']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        # Time log data
        row = 17
        total_hours = 0
        for _, log_row in group.iterrows():
            time_in = pd.to_datetime(log_row['time_in']) if pd.notna(log_row['time_in']) else None
            time_out = pd.to_datetime(log_row['time_out']) if pd.notna(log_row['time_out']) else None

            ws.cell(row=row, column=1, value=time_in.strftime('%m/%d/%Y') if time_in is not None else 'MISSING TIME IN')
            ws.cell(row=row, column=2, value=time_in.strftime('%I:%M %p') if time_in is not None else 'MISSING')
            ws.cell(row=row, column=3, value=time_out.strftime('%I:%M %p') if time_out is not None else 'INCOMPLETE TICKET')
            ws.cell(row=row, column=4, value=round(log_row['hours'], 2) if pd.notna(log_row['hours']) else 0)
            ws.cell(row=row, column=5, value=log_row['contractor_name'] or '')
            ws.cell(row=row, column=6, value=log_row['worker_name'] or '')
            ws.cell(row=row, column=7, value='Snow Removal')
            ws.cell(row=row, column=8, value=log_row['notes'] or '')
            total_hours += log_row['hours'] if pd.notna(log_row['hours']) else 0
            row += 1

        # Total row
        ws.cell(row=row, column=2, value='TOTAL:').font = Font(bold=True)
        ws.cell(row=row, column=3, value=round(total_hours, 2)).font = Font(bold=True)

        # Weather Report Section
        row += 3
        ws.cell(row=row, column=1, value='24-Hour Weather Report').font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        # Fetch weather data if API key available
        weather_api_key = get_weather_api_key()
        logger.info(f"Generating weather report for property {first_row.get('property_name', 'Unknown')} - API key: {'Found' if weather_api_key else 'Missing'}, Coordinates: {first_row.get('latitude')}, {first_row.get('longitude')}")
        if weather_api_key and first_row['latitude'] and first_row['longitude']:
            try:
                import requests
                from datetime import datetime, timedelta

                # Get earliest time_in from this group for weather data
                min_time = group['time_in'].min()
                event_timestamp = int(pd.to_datetime(min_time).timestamp())

                lat = first_row['latitude']
                lon = first_row['longitude']

                # Use OpenWeather One Call API 3.0 timemachine for historical data
                # This endpoint provides historical weather for any past date
                weather_url = f"https://api.openweathermap.org/data/3.0/onecall/timemachine"
                params = {
                    'lat': lat,
                    'lon': lon,
                    'dt': event_timestamp,
                    'appid': weather_api_key,
                    'units': 'imperial'
                }

                logger.info(f"Fetching weather for event at timestamp {event_timestamp} ({pd.to_datetime(min_time)})")

                response = requests.get(weather_url, params=params, timeout=10)

                logger.info(f"Weather API request to {weather_url} - Status: {response.status_code}")
                logger.debug(f"Weather API Response: {response.text[:500]}")

                if response.status_code == 200:
                    weather_data = response.json()
                    logger.info(f"Weather data keys: {weather_data.keys()}")

                    # One Call API 3.0 timemachine returns data in 'data' array
                    # Each hour has its own entry, we'll use the first one (closest to event time)
                    if 'data' in weather_data and len(weather_data['data']) > 0:
                        # Timemachine API returns hourly data array
                        hourly_data = weather_data['data'][0]

                        temp = hourly_data.get('temp', 'N/A')
                        feels_like = hourly_data.get('feels_like', 'N/A')
                        conditions = hourly_data.get('weather', [{}])[0].get('description', 'Unknown').capitalize()
                        wind_speed = hourly_data.get('wind_speed', 0)
                        humidity = hourly_data.get('humidity', 'N/A')

                        # Check for precipitation (1h totals)
                        rain = hourly_data.get('rain', {}).get('1h', 0) if isinstance(hourly_data.get('rain'), dict) else hourly_data.get('rain', 0)
                        snow = hourly_data.get('snow', {}).get('1h', 0) if isinstance(hourly_data.get('snow'), dict) else hourly_data.get('snow', 0)
                    else:
                        # Fallback - shouldn't happen with correct API
                        logger.warning(f"Unexpected weather data format: {weather_data}")
                        temp = 'N/A'
                        feels_like = 'N/A'
                        conditions = 'Data unavailable'
                        wind_speed = 0
                        humidity = 'N/A'
                        rain = 0
                        snow = 0

                    row += 1
                    ws.cell(row=row, column=1, value='Conditions:')
                    ws.cell(row=row, column=2, value=conditions)

                    row += 1
                    ws.cell(row=row, column=1, value='Temperature:')
                    ws.cell(row=row, column=2, value=f"{temp}°F (Feels like {feels_like}°F)")

                    row += 1
                    ws.cell(row=row, column=1, value='Wind Speed:')
                    ws.cell(row=row, column=2, value=f"{wind_speed} mph")

                    row += 1
                    ws.cell(row=row, column=1, value='Humidity:')
                    ws.cell(row=row, column=2, value=f"{humidity}%")

                    if snow > 0:
                        row += 1
                        ws.cell(row=row, column=1, value='Snow (1hr):')
                        ws.cell(row=row, column=2, value=f"{snow} inches")

                    if rain > 0:
                        row += 1
                        ws.cell(row=row, column=1, value='Rain (1hr):')
                        ws.cell(row=row, column=2, value=f"{rain} inches")
                else:
                    logger.error(f"Weather API failed: Status {response.status_code}, Response: {response.text}")
                    row += 1
                    ws.cell(row=row, column=1, value='Weather data:')
                    ws.cell(row=row, column=2, value=f'Unable to fetch (API Status: {response.status_code})')

            except Exception as e:
                logger.error(f"Weather fetch error: {e}", exc_info=True)
                row += 1
                ws.cell(row=row, column=1, value='Weather data:')
                ws.cell(row=row, column=2, value='Contact area manager for details')
        else:
            row += 1
            ws.cell(row=row, column=1, value='Weather data:')
            ws.cell(row=row, column=2, value='API key not configured or no location data')

        # Adjust column widths
        ws.column_dimensions['A'].width = 25.14  # Date
        ws.column_dimensions['B'].width = 15  # Time In
        ws.column_dimensions['C'].width = 15  # Time Out
        ws.column_dimensions['D'].width = 12  # Total Hours
        ws.column_dimensions['E'].width = 20  # Contractor
        ws.column_dimensions['F'].width = 20  # Worker Name
        ws.column_dimensions['G'].width = 20  # Service Performed
        ws.column_dimensions['H'].width = 30  # Notes

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"field_service_reports_{start}_{end}.xlsx" if start and end else "field_service_reports.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
