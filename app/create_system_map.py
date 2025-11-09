import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create Excel writer
output_file = 'ContractorPortal_SystemMap.xlsx'
writer = pd.ExcelWriter(output_file, engine='openpyxl')

# Sheet 1: Database Tables Schema
db_tables_data = [
    ['Table Name', 'Column Name', 'Data Type', 'Constraints', 'Description'],
    ['users', 'id', 'INT', 'PRIMARY KEY, AUTO_INCREMENT', 'Unique user identifier'],
    ['users', 'name', 'VARCHAR(255)', 'NOT NULL', 'User full name (Last, First for subcontractors)'],
    ['users', 'phone', 'VARCHAR(20)', '', 'Phone number'],
    ['users', 'username', 'VARCHAR(100)', 'UNIQUE', 'Login username'],
    ['users', 'email', 'VARCHAR(255)', 'UNIQUE, NOT NULL', 'Email address'],
    ['users', 'role', 'VARCHAR(50)', 'NOT NULL', 'Admin, Manager, or Subcontractor'],
    ['users', 'contractor_id', 'INT', 'FOREIGN KEY -> users(id)', 'Links workers to their contractor/manager'],
    ['users', 'password', 'VARCHAR(255)', '', 'Legacy password field'],
    ['users', 'password_hash', 'VARCHAR(255)', '', 'Hashed password'],
    ['users', 'status', 'VARCHAR(20)', "DEFAULT 'pending'", 'pending, active, suspended'],
    ['users', 'created_at', 'TIMESTAMP', 'DEFAULT CURRENT_TIMESTAMP', 'Account creation date'],
    ['users', 'updated_at', 'TIMESTAMP', 'ON UPDATE CURRENT_TIMESTAMP', 'Last update date'],
    ['', '', '', '', ''],
    ['locations', 'id', 'INT', 'PRIMARY KEY, AUTO_INCREMENT', 'Unique property identifier'],
    ['locations', 'name', 'VARCHAR(255)', 'NOT NULL', 'Property name'],
    ['locations', 'address', 'VARCHAR(255)', 'NOT NULL, UNIQUE', 'Property address'],
    ['locations', 'sqft', 'INT', '', 'Lot square footage'],
    ['locations', 'area_manager', 'VARCHAR(100)', '', 'Assigned manager name'],
    ['locations', 'plow', 'TINYINT(1)', 'DEFAULT 0', 'Property requires plowing service'],
    ['locations', 'salt', 'TINYINT(1)', 'DEFAULT 0', 'Property requires salting service'],
    ['locations', 'notes', 'TEXT', '', 'Additional notes'],
    ['locations', 'created_at', 'TIMESTAMP', 'DEFAULT CURRENT_TIMESTAMP', 'Record creation date'],
    ['locations', 'updated_at', 'TIMESTAMP', 'ON UPDATE CURRENT_TIMESTAMP', 'Last update date'],
    ['', '', '', '', ''],
    ['winter_ops_logs', 'id', 'INT', 'PRIMARY KEY, AUTO_INCREMENT', 'Unique log entry identifier'],
    ['winter_ops_logs', 'property_id', 'INT', 'FOREIGN KEY -> locations(id)', 'Property where work was performed'],
    ['winter_ops_logs', 'contractor_id', 'INT', 'FOREIGN KEY -> users(id)', 'Contractor/Manager ID'],
    ['winter_ops_logs', 'contractor_name', 'VARCHAR(255)', '', 'Contractor/Manager company name'],
    ['winter_ops_logs', 'worker_name', 'VARCHAR(255)', 'NOT NULL', 'Individual worker name (Last, First)'],
    ['winter_ops_logs', 'equipment', 'VARCHAR(255)', '', 'Equipment used for the job'],
    ['winter_ops_logs', 'time_in', 'DATETIME', 'NOT NULL', 'Work start time'],
    ['winter_ops_logs', 'time_out', 'DATETIME', 'NOT NULL', 'Work end time'],
    ['winter_ops_logs', 'bulk_salt_qty', 'DECIMAL(10,2)', 'DEFAULT 0', 'Bulk salt quantity (yards)'],
    ['winter_ops_logs', 'bag_salt_qty', 'DECIMAL(10,2)', 'DEFAULT 0', 'Bagged salt quantity'],
    ['winter_ops_logs', 'calcium_chloride_qty', 'DECIMAL(10,2)', 'DEFAULT 0', 'Calcium chloride quantity (bags)'],
    ['winter_ops_logs', 'customer_provided', 'TINYINT(1)', 'DEFAULT 0', 'Whether materials were customer provided'],
    ['winter_ops_logs', 'notes', 'TEXT', '', 'Additional notes about the work'],
    ['winter_ops_logs', 'created_at', 'TIMESTAMP', 'DEFAULT CURRENT_TIMESTAMP', 'Record creation date'],
    ['winter_ops_logs', 'updated_at', 'TIMESTAMP', 'ON UPDATE CURRENT_TIMESTAMP', 'Last update date'],
    ['', '', '', '', ''],
    ['green_services_logs', 'id', 'INT', 'PRIMARY KEY, AUTO_INCREMENT', 'Unique log entry identifier'],
    ['green_services_logs', 'property_id', 'INT', 'FOREIGN KEY -> locations(id)', 'Property where work was performed'],
    ['green_services_logs', 'subcontractor_name', 'VARCHAR(255)', 'NOT NULL', 'Worker name'],
    ['green_services_logs', 'time_in', 'DATETIME', 'NOT NULL', 'Work start time'],
    ['green_services_logs', 'time_out', 'DATETIME', 'NOT NULL', 'Work end time'],
    ['green_services_logs', 'service_type', 'VARCHAR(100)', 'NOT NULL', 'Type of service (mowing, trimming, etc)'],
    ['green_services_logs', 'products_used', 'VARCHAR(255)', '', 'Products/materials used'],
    ['green_services_logs', 'quantity_used', 'DECIMAL(10,2)', 'DEFAULT 0', 'Quantity of products used'],
    ['green_services_logs', 'notes', 'TEXT', '', 'Additional notes'],
    ['green_services_logs', 'created_at', 'TIMESTAMP', 'DEFAULT CURRENT_TIMESTAMP', 'Record creation date'],
    ['green_services_logs', 'updated_at', 'TIMESTAMP', 'ON UPDATE CURRENT_TIMESTAMP', 'Last update date'],
    ['', '', '', '', ''],
    ['equipment_rates', 'id', 'INT', 'PRIMARY KEY, AUTO_INCREMENT', 'Unique equipment identifier'],
    ['equipment_rates', 'equipment_name', 'VARCHAR(255)', 'NOT NULL, UNIQUE', 'Equipment type name'],
    ['equipment_rates', 'hourly_rate', 'DECIMAL(10,2)', 'NOT NULL, DEFAULT 0', 'Hourly rate for equipment'],
    ['equipment_rates', 'description', 'TEXT', '', 'Equipment description'],
    ['equipment_rates', 'created_at', 'TIMESTAMP', 'DEFAULT CURRENT_TIMESTAMP', 'Record creation date'],
    ['equipment_rates', 'updated_at', 'TIMESTAMP', 'ON UPDATE CURRENT_TIMESTAMP', 'Last update date'],
    ['', '', '', '', ''],
    ['winter_products', 'id', 'INT', 'PRIMARY KEY, AUTO_INCREMENT', 'Unique product identifier'],
    ['winter_products', 'name', 'VARCHAR(255)', 'NOT NULL', 'Product name'],
    ['winter_products', 'unit', 'VARCHAR(50)', 'NOT NULL', 'Unit of measurement'],
    ['winter_products', 'created_at', 'TIMESTAMP', 'DEFAULT CURRENT_TIMESTAMP', 'Record creation date'],
    ['winter_products', 'updated_at', 'TIMESTAMP', 'ON UPDATE CURRENT_TIMESTAMP', 'Last update date'],
    ['', '', '', '', ''],
    ['landscape_products', 'id', 'INT', 'PRIMARY KEY, AUTO_INCREMENT', 'Unique product identifier'],
    ['landscape_products', 'name', 'VARCHAR(255)', 'NOT NULL', 'Product name'],
    ['landscape_products', 'unit', 'VARCHAR(50)', 'NOT NULL', 'Unit of measurement'],
    ['landscape_products', 'created_at', 'TIMESTAMP', 'DEFAULT CURRENT_TIMESTAMP', 'Record creation date'],
    ['landscape_products', 'updated_at', 'TIMESTAMP', 'ON UPDATE CURRENT_TIMESTAMP', 'Last update date'],
    ['', '', '', '', ''],
    ['oauth_identities', 'id', 'INT', 'PRIMARY KEY, AUTO_INCREMENT', 'Unique identity identifier'],
    ['oauth_identities', 'user_id', 'INT', 'FOREIGN KEY -> users(id)', 'Linked user account'],
    ['oauth_identities', 'provider', 'VARCHAR(50)', 'NOT NULL', 'OAuth provider (google, microsoft)'],
    ['oauth_identities', 'provider_user_id', 'VARCHAR(255)', 'NOT NULL', 'Provider-specific user ID'],
    ['oauth_identities', 'email', 'VARCHAR(255)', 'NOT NULL', 'Email from OAuth provider'],
    ['oauth_identities', 'created_at', 'TIMESTAMP', 'DEFAULT CURRENT_TIMESTAMP', 'Record creation date'],
]

df_db = pd.DataFrame(db_tables_data[1:], columns=db_tables_data[0])
df_db.to_excel(writer, sheet_name='Database Schema', index=False)

print('Created Database Schema sheet')

# Sheet 2: API Endpoints
api_data = [
    ['Endpoint', 'Method', 'Purpose', 'Database Tables', 'Required Role', 'Route File'],
    ['/api/login/', 'POST', 'User authentication', 'users', 'Public', 'auth_routes.py'],
    ['/auth/google/start', 'GET', 'Google SSO redirect', 'users, oauth_identities', 'Public', 'routers/auth_oidc.py'],
    ['/auth/microsoft/start', 'GET', 'Microsoft SSO redirect', 'users, oauth_identities', 'Public', 'routers/auth_oidc.py'],
    ['/signup-request/', 'POST', 'New user signup', 'users', 'Public', 'misc_routes.py'],
    ['/users/', 'GET', 'Get all users', 'users', 'Admin', 'auth_routes.py'],
    ['/contractors/', 'GET', 'Get all contractors/managers', 'users', 'All', 'auth_routes.py'],
    ['/add-user/', 'POST', 'Add new user', 'users', 'Admin', 'auth_routes.py'],
    ['/update-user/{id}', 'PUT', 'Update user', 'users', 'Admin', 'auth_routes.py'],
    ['/delete-user/{id}', 'DELETE', 'Delete user', 'users', 'Admin', 'auth_routes.py'],
    ['/reset-password/', 'PUT', 'Reset user password', 'users', 'Admin', 'misc_routes.py'],
    ['/admin/pending-users/', 'GET', 'Get pending signups', 'users', 'Admin', 'auth_routes.py'],
    ['/admin/approve-user/{id}', 'POST', 'Approve user signup', 'users', 'Admin', 'auth_routes.py'],
    ['/admin/reject-user/{id}', 'POST', 'Reject user signup', 'users', 'Admin', 'auth_routes.py'],
    ['/properties/', 'GET', 'Get all properties', 'locations', 'All', 'property_routes.py'],
    ['/add-property/', 'POST', 'Add new property', 'locations', 'Admin/Manager', 'property_routes.py'],
    ['/update-property/', 'PUT', 'Update property', 'locations', 'Admin/Manager', 'property_routes.py'],
    ['/delete-property/{id}', 'DELETE', 'Delete property', 'locations', 'Admin/Manager', 'property_routes.py'],
    ['/bulk-import-properties/', 'POST', 'Bulk import from Excel', 'locations', 'Admin/Manager', 'property_routes.py'],
    ['/submit-winter-log/', 'POST', 'Submit winter ops log', 'winter_ops_logs', 'All', 'ops_routes.py'],
    ['/winter-logs/', 'GET', 'Get winter ops logs', 'winter_ops_logs, locations', 'All', 'ops_routes.py'],
    ['/submit-green-log/', 'POST', 'Submit green services log', 'green_services_logs', 'All', 'ops_routes.py'],
    ['/green-logs/', 'GET', 'Get green services logs', 'green_services_logs, locations', 'All', 'ops_routes.py'],
    ['/equipment-rates/', 'GET', 'Get equipment rates', 'equipment_rates', 'All', 'equipment_routes.py'],
    ['/equipment-rates/', 'POST', 'Add equipment rate', 'equipment_rates', 'Admin/Manager', 'equipment_routes.py'],
    ['/equipment-rates/{id}', 'PUT', 'Update equipment rate', 'equipment_rates', 'Admin/Manager', 'equipment_routes.py'],
    ['/equipment-rates/{id}', 'DELETE', 'Delete equipment rate', 'equipment_rates', 'Admin', 'equipment_routes.py'],
    ['/equipment-usage-report/', 'GET', 'Equipment usage report', 'winter_ops_logs, equipment_rates', 'All', 'equipment_routes.py'],
    ['/winter-products/', 'GET', 'Get winter products', 'winter_products', 'All', 'product_routes.py'],
    ['/winter-products/', 'POST', 'Add winter product', 'winter_products', 'Admin/Manager', 'product_routes.py'],
    ['/winter-products/{id}', 'PUT', 'Update winter product', 'winter_products', 'Admin/Manager', 'product_routes.py'],
    ['/winter-products/{id}', 'DELETE', 'Delete winter product', 'winter_products', 'Admin/Manager', 'product_routes.py'],
    ['/landscape-products/', 'GET', 'Get landscape products', 'landscape_products', 'All', 'product_routes.py'],
    ['/landscape-products/', 'POST', 'Add landscape product', 'landscape_products', 'Admin', 'product_routes.py'],
    ['/landscape-products/{id}', 'PUT', 'Update landscape product', 'landscape_products', 'Admin', 'product_routes.py'],
    ['/landscape-products/{id}', 'DELETE', 'Delete landscape product', 'landscape_products', 'Admin', 'product_routes.py'],
    ['/export/contractor-timesheets/', 'POST', 'Export contractor timesheets', 'winter_ops_logs, locations', 'All', 'report_routes.py'],
    ['/export/property-logs/', 'POST', 'Export property logs', 'winter_ops_logs, locations', 'All', 'report_routes.py'],
]

df_api = pd.DataFrame(api_data[1:], columns=api_data[0])
df_api.to_excel(writer, sheet_name='API Endpoints', index=False)

print('Created API Endpoints sheet')

# Sheet 3: Page to Database Mapping
page_db_data = [
    ['HTML Page', 'Purpose', 'Database Tables (Read)', 'Database Tables (Write)', 'Roles'],
    ['login.html', 'User authentication', 'users', '', 'Public'],
    ['signup.html', 'New user registration', '', 'users', 'Public'],
    ['AdminDashboard.html', 'Admin main dashboard', 'users', 'users', 'Admin'],
    ['ApprovePendingUsers.html', 'Approve new user signups', 'users', 'users', 'Admin'],
    ['ManagerDashboard.html', 'Manager main dashboard', '', '', 'Manager, Admin'],
    ['UserDashboard.html', 'Subcontractor main dashboard', '', '', 'Subcontractor'],
    ['ManageUsers.html', 'User CRUD operations', 'users', 'users', 'Admin'],
    ['PropertyInfo.html', 'Property CRUD and bulk import', 'locations', 'locations', 'All'],
    ['SnowRemoval.html', 'Submit winter ops logs', 'locations, users, equipment_rates', 'winter_ops_logs', 'All'],
    ['WinterOpsLog.html', 'Submit winter ops logs (duplicate)', 'locations, users, equipment_rates', 'winter_ops_logs', 'All'],
    ['ViewWinterLogs.html', 'View/filter winter logs', 'winter_ops_logs, locations', '', 'All'],
    ['WinterOpsAdmin.html', 'Admin view winter logs', 'winter_ops_logs, locations', '', 'All'],
    ['LawnMowing.html', 'Submit green services logs', 'locations', 'green_services_logs', 'All'],
    ['ViewGreenLogs.html', 'View/filter green logs', 'green_services_logs, locations', '', 'All'],
    ['GreenOpsAdmin.html', 'Admin view green logs', 'green_services_logs, locations', '', 'All'],
    ['Reports.html', 'Generate reports and export Excel', 'winter_ops_logs, locations', '', 'All'],
    ['ManageWinterProducts.html', 'Manage winter products', 'winter_products', 'winter_products', 'Admin, Manager'],
    ['ManageLandscapeProducts.html', 'Manage landscape products', 'landscape_products', 'landscape_products', 'Admin'],
]

df_page_db = pd.DataFrame(page_db_data[1:], columns=page_db_data[0])
df_page_db.to_excel(writer, sheet_name='Page-Database Map', index=False)

print('Created Page-Database Map sheet')

# Sheet 4: Role-Based Access Matrix
role_data = [
    ['Feature/Page', 'Admin', 'Manager', 'Subcontractor', 'Notes'],
    ['Login', 'Yes', 'Yes', 'Yes', 'Public access'],
    ['Signup', 'Yes', 'Yes', 'Yes', 'Public access'],
    ['Admin Dashboard', 'Yes', 'No', 'No', 'Admin only'],
    ['Manager Dashboard', 'Yes', 'Yes', 'No', 'Managers and Admins'],
    ['User Dashboard', 'No', 'No', 'Yes', 'Subcontractors only'],
    ['Manage Users', 'Yes', 'No', 'No', 'Full CRUD - Admin only'],
    ['Approve Pending Users', 'Yes', 'No', 'No', 'Admin only'],
    ['View Properties', 'Yes', 'Yes', 'Yes', 'All roles'],
    ['Add/Edit Properties', 'Yes', 'Yes', 'No', 'Admin and Manager'],
    ['Delete Properties', 'Yes', 'Yes', 'No', 'Admin and Manager'],
    ['Bulk Import Properties', 'Yes', 'Yes', 'No', 'Admin and Manager'],
    ['Submit Winter Logs', 'Yes', 'Yes', 'Yes', 'All roles'],
    ['View Winter Logs', 'Yes', 'Yes', 'Yes', 'All roles'],
    ['Submit Green Services Logs', 'Yes', 'Yes', 'Yes', 'All roles'],
    ['View Green Services Logs', 'Yes', 'Yes', 'Yes', 'All roles'],
    ['Manage Equipment Rates', 'Yes', 'Yes', 'No', 'Admin and Manager'],
    ['View Equipment Rates', 'Yes', 'Yes', 'Yes', 'All roles'],
    ['Delete Equipment Rates', 'Yes', 'No', 'No', 'Admin only'],
    ['Manage Winter Products', 'Yes', 'Yes', 'No', 'Admin and Manager'],
    ['Manage Landscape Products', 'Yes', 'No', 'No', 'Admin only'],
    ['Generate Reports', 'Yes', 'Yes', 'Yes', 'All roles'],
    ['Export to Excel', 'Yes', 'Yes', 'Yes', 'All roles'],
]

df_roles = pd.DataFrame(role_data[1:], columns=role_data[0])
df_roles.to_excel(writer, sheet_name='Role Access Matrix', index=False)

print('Created Role Access Matrix sheet')

# Sheet 5: Issues and Missing Connections
issues_data = [
    ['Issue #', 'Type', 'Severity', 'Description', 'Affected Pages/Endpoints', 'Current Status'],
    [1, 'Inconsistency', 'Medium', 'green_services_logs still uses subcontractor_name, not updated to contractor/worker structure', 'LawnMowing.html, GreenOpsLog.html, ViewGreenLogs.html', 'NOT FIXED'],
    [2, 'Missing Feature', 'Low', 'No Equipment Management UI page for Admin/Manager to manage equipment_rates', 'None - missing page', 'NOT IMPLEMENTED'],
    [3, 'Missing Feature', 'Low', 'No Equipment Usage Report UI page for viewing equipment hours and costs', 'None - missing page', 'NOT IMPLEMENTED'],
    [4, 'Inconsistency', 'Low', 'Reports.html Excel export still uses subcontractor_name field instead of contractor_name/worker_name', 'Reports.html, report_routes.py', 'NOT FIXED'],
    [5, 'Data Structure', 'Medium', 'ManageUsers.html does not allow assigning contractor_id to workers', 'ManageUsers.html', 'NOT FIXED'],
    [6, 'Missing Relationship', 'Low', 'green_services_logs does not track contractor_id or link to users table', 'green_services_logs table, ops_routes.py', 'NOT FIXED'],
    [7, 'Report Issue', 'Medium', 'report_routes.py queries still reference old subcontractor_name instead of worker_name', 'report_routes.py (/export endpoints)', 'NOT FIXED'],
    [8, 'Missing Feature', 'Low', 'No admin page to view/manage OAuth identities', 'None - missing page', 'NOT IMPLEMENTED'],
    [9, 'Data Validation', 'Low', 'No validation for Last, First name format for workers', 'SnowRemoval.html, WinterOpsLog.html', 'NOT IMPLEMENTED'],
    [10, 'Duplicate Files', 'Low', 'WinterOpsLog.html and SnowRemoval.html are duplicates', 'WinterOpsLog.html, SnowRemoval.html', 'BOTH EXIST'],
    [11, 'Missing Link', 'Low', 'No navigation from dashboards to Equipment Management (if implemented)', 'Admin/Manager Dashboards', 'NOT LINKED'],
    [12, 'Missing Link', 'Low', 'No navigation from dashboards to ViewWinterLogs.html or ViewGreenLogs.html', 'All dashboards', 'NOT LINKED'],
    [13, 'Data Integrity', 'Medium', 'equipment field in winter_ops_logs is VARCHAR, not FK to equipment_rates', 'winter_ops_logs table', 'BY DESIGN (allows flexibility)'],
    [14, 'Missing Feature', 'Low', 'No way to edit or delete winter_ops_logs or green_services_logs after submission', 'ViewWinterLogs.html, ViewGreenLogs.html', 'NOT IMPLEMENTED'],
    [15, 'Report Issue', 'Medium', 'Report queries use subcontractor_name joins that will fail with new structure', 'report_routes.py (all report endpoints)', 'NOT FIXED'],
]

df_issues = pd.DataFrame(issues_data[1:], columns=issues_data[0])
df_issues.to_excel(writer, sheet_name='Issues & Missing Items', index=False)

print('Created Issues & Missing Items sheet')

# Sheet 6: Navigation Flow
nav_data = [
    ['From Page', 'Link Text/Action', 'To Page', 'Role Required', 'Notes'],
    ['Home.html', 'Login', 'login.html', 'Public', 'Homepage link'],
    ['Home.html', 'Sign Up', 'signup.html', 'Public', 'Homepage link'],
    ['login.html', 'OAuth Login Success', 'AdminDashboard.html', 'Admin', 'Role-based redirect'],
    ['login.html', 'OAuth Login Success', 'ManagerDashboard.html', 'Manager', 'Role-based redirect'],
    ['login.html', 'OAuth Login Success', 'UserDashboard.html', 'Subcontractor', 'Role-based redirect'],
    ['signup.html', 'Already have account', 'login.html', 'Public', 'Link'],
    ['signup.html', 'OAuth Signup Success', 'pending-approval.html', 'Public', 'Awaiting admin approval'],
    ['pending-approval.html', 'Back to Login', 'login.html', 'Public', 'Link'],
    ['AdminDashboard.html', 'Manage Properties', 'PropertyInfo.html', 'Admin', 'Navigation link'],
    ['AdminDashboard.html', 'Manage Users', 'ManageUsers.html', 'Admin', 'Navigation link'],
    ['AdminDashboard.html', 'Generate Reports', 'Reports.html', 'Admin', 'Navigation link'],
    ['AdminDashboard.html', 'Winter Products', 'ManageWinterProducts.html', 'Admin', 'Navigation link'],
    ['AdminDashboard.html', 'Landscape Products', 'ManageLandscapeProducts.html', 'Admin', 'Navigation link'],
    ['AdminDashboard.html', 'Submit Winter Log', 'WinterOpsLog.html', 'Admin', 'Navigation link'],
    ['AdminDashboard.html', 'Submit Green Log', 'GreenOpsLog.html', 'Admin', 'Navigation link'],
    ['ManagerDashboard.html', 'View Properties', 'PropertyInfo.html', 'Manager', 'Navigation link'],
    ['ManagerDashboard.html', 'Generate Reports', 'Reports.html', 'Manager', 'Navigation link'],
    ['ManagerDashboard.html', 'Winter Operations Log', 'WinterOpsLog.html', 'Manager', 'Navigation link'],
    ['ManagerDashboard.html', 'Green Operations Log', 'GreenOpsLog.html', 'Manager', 'Navigation link'],
    ['ManagerDashboard.html', 'Return Home', 'Home.html', 'Manager', 'Navigation link'],
    ['UserDashboard.html', 'View Properties', 'PropertyInfo.html', 'Subcontractor', 'Navigation link'],
    ['UserDashboard.html', 'Submit Winter Log', 'WinterOpsLog.html', 'Subcontractor', 'Navigation link'],
    ['UserDashboard.html', 'Submit Green Log', 'GreenOpsLog.html', 'Subcontractor', 'Navigation link'],
]

df_nav = pd.DataFrame(nav_data[1:], columns=nav_data[0])
df_nav.to_excel(writer, sheet_name='Navigation Flow', index=False)

print('Created Navigation Flow sheet')

# Save and close
writer.close()

# Apply formatting
wb = openpyxl.load_workbook(output_file)

# Format each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Header formatting
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze top row
    ws.freeze_panes = 'A2'

wb.save(output_file)

print(f'\nExcel file successfully created: {output_file}')
print('All sheets formatted with headers and column widths adjusted')
